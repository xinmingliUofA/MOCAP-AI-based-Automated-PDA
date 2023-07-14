import pandas as pd
import os

from openpyxl import load_workbook


def load_data(file_name):
    data_dir = 'data'
    file_dir = os.path.join(data_dir, file_name)
    data = pd.read_excel(file_dir, sheet_name=['Segment Position', 'Joint Angles XZY', 'Ergonomic Joint Angles XZY'])

    return pd.merge(pd.merge(data['Ergonomic Joint Angles XZY'], data['Segment Position'], on='Frame'),
                    data['Joint Angles XZY'], on='Frame').iloc[1:, :]


# def concat_data(file1, file2):
#     data1 = load_data(file1)
#     data2 = load_data(file2)
#     data2['Frame'] += data1['Frame'].iloc[-1]
#     return pd.concat([data1, data2], ignore_index=True)
#
# file1 = 'Test-1.xlsx'
# file2 = 'Test-2.xlsx'
# data = concat_data(file1, file2)
# data.to_excel('merged.xlsx')

file_name = 'Test.xlsx'
data = load_data(file_name)

total_time = data['Frame'].iloc[-1] / 60

# Organs height

Knee_height = data.iloc[0]['Right Lower Leg z']
Waist_height = data.iloc[0]['Right Upper Leg z']
Shoulder_height = data.iloc[0]['Right Shoulder z']
A = (Waist_height - Knee_height)
B = (Shoulder_height - Waist_height)

# Low Level Lifting

lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) >= 0.01]
not_lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) < 0.01]
lifting_lz = not_lifting_rz[(not_lifting_rz['Left Hand z'] - not_lifting_rz['Left Hand z'].shift()).round(2) >= 0.01]

low_level_lifting_data_rz = lifting_rz.query(
    '(`Right Hand z` < `Right Lower Leg z` | `Left Hand z`  < `Left Lower Leg z`) &'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

low_level_lifting_data_lz = lifting_lz.query(
    '(`Right Hand z` < `Right Lower Leg z` | `Left Hand z` < `Left Lower Leg z`) &'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

low_level_lifting = len(low_level_lifting_data_rz) + len(low_level_lifting_data_lz)

total_time_low_level_lifting = low_level_lifting / 6

low_level_lifting_freq = total_time_low_level_lifting / total_time * 100

print('low_level_lifting_freq', round(low_level_lifting_freq, 2))
# print(low_level_lifting_data_rz)
# print(low_level_lifting_data_lz)

# ########################################################

# Knee Level Lifting

lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) >= 0.01]
not_lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) < 0.01]
lifting_lz = not_lifting_rz[(not_lifting_rz['Left Hand z'] - not_lifting_rz['Left Hand z'].shift()).round(2) >= 0.01]

Knee_level_lifting_data_rz = lifting_rz.query(
    '`Vertical_Pelvis Flexion/Extension` >= 20 & '
    ' (`Right Lower Leg z` <= `Right Hand z` < `Right Lower Leg z` + @A/2) |'
    ' (`Left Lower Leg z` <= `Left Hand z` < `Left Lower Leg z` + @A/2) &'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(`Right Ankle Dorsiflexion/Plantarflexion` < 15 & `Left Ankle Dorsiflexion/Plantarflexion` < 15) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

Knee_level_lifting_data_lz = lifting_lz.query(
    '`Vertical_Pelvis Flexion/Extension` >= 20 & '
    '(`Right Lower Leg z` <= `Right Hand z` < `Right Lower Leg z` + @A/2) |'
    '(`Left Lower Leg z` <= `Left Hand z` < `Left Lower Leg z` + @A/2) &'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(`Right Ankle Dorsiflexion/Plantarflexion` < 15 & `Left Ankle Dorsiflexion/Plantarflexion` < 15) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

Knee_level_lifting = len(Knee_level_lifting_data_rz) + len(Knee_level_lifting_data_lz)

total_time_Knee_level_lifting = Knee_level_lifting / 6

Knee_level_lifting_freq = total_time_Knee_level_lifting / total_time * 100

print('Knee_level_lifting_freq', round(Knee_level_lifting_freq, 2))
# print(Knee_level_lifting_data_rz)
# print(Knee_level_lifting_data_lz)

# ########################################################

# Waist Level Lifting

lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) >= 0.01]
not_lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) < 0.01]
lifting_lz = not_lifting_rz[(not_lifting_rz['Left Hand z'] - not_lifting_rz['Left Hand z'].shift()).round(2) >= 0.01]

waist_level_lifting_data_rz = lifting_rz.query(
    '((`Right Lower Leg z` + @A/2 <= `Right Hand z` < `Pelvis z` + @B/2) |'
    '(`Left Lower Leg z` + @A/2 <= `Left Hand z` < `Pelvis z` + @B/2)) &'
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(`Right Shoulder Abduction/Adduction` < 30 & `Left Shoulder Abduction/Adduction` < 30) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

waist_level_lifting_data_lz = lifting_lz.query(
    '((`Right Lower Leg z` + @A/2 <= `Right Hand z` < `Pelvis z` + @B/2) |'
    '(`Left Lower Leg z` + @A/2 <= `Left Hand z` < `Pelvis z` + @B/2)) &'
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(`Right Shoulder Abduction/Adduction` < 30 & `Left Shoulder Abduction/Adduction` < 30) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

waist_level_lifting = len(waist_level_lifting_data_rz) + len(waist_level_lifting_data_lz)

total_time_waist_level_lifting = waist_level_lifting / 6

waist_level_lifting_freq = total_time_waist_level_lifting / total_time * 100

print('waist_level_lifting_freq', round(waist_level_lifting_freq, 2))
# print(waist_level_lifting_data_rz)
# print(waist_level_lifting_data_lz)

# ####################################################

# Shoulder Level Lifting

lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) >= 0.01]
not_lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) < 0.01]
lifting_lz = not_lifting_rz[(not_lifting_rz['Left Hand z'] - not_lifting_rz['Left Hand z'].shift()).round(2) >= 0.01]

shoulder_level_lifting_data_rz = lifting_rz.query(
    '(`Pelvis z` + @B/2 <= `Right Hand z` < `Right Shoulder z` + 0.1 |'
    ' `Pelvis z` + @B/2 <= `Left Hand z` < `Left Shoulder z` + 0.1) &'
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

shoulder_level_lifting_data_lz = lifting_lz.query(
    '(`Pelvis z` + @B/2 <= `Right Hand z` < `Right Shoulder z` + 0.1 |'
    ' `Pelvis z` + @B/2 <= `Left Hand z` < `Left Shoulder z` + 0.1) &'
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

Shoulder_level_lifting = len(shoulder_level_lifting_data_rz) + len(shoulder_level_lifting_data_lz)

total_time_Shoulder_level_lifting = Shoulder_level_lifting / 6

Shoulder_level_lifting_freq = total_time_Shoulder_level_lifting / total_time * 100

print('Shoulder_level_lifting_freq', round(Shoulder_level_lifting_freq, 2))
# print(shoulder_level_lifting_data_rz)
# print(shoulder_level_lifting_data_lz)

# ####################################################

# Above Shoulder Level Lifting

lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) >= 0.01]
not_lifting_rz = data[(data['Right Hand z'] - data['Right Hand z'].shift()).round(2) < 0.01]
lifting_lz = not_lifting_rz[(not_lifting_rz['Left Hand z'] - not_lifting_rz['Left Hand z'].shift()).round(2) >= 0.01]

Above_shoulder_level_lifting_data_rz = lifting_rz.query(
    '(`Right Hand z` >= `Right Shoulder z` + 0.1 | `Left Hand z` >= `Left Shoulder z` + 0.1)'
    '& (`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

Above_shoulder_level_lifting_data_lz = lifting_lz.query(
    '(`Right Hand z` >= `Right Shoulder z` + 0.1 | `Left Hand z` >= `Left Shoulder z` + 0.1)'
    '& (`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) ')

Above_Shoulder_level_lifting = len(Above_shoulder_level_lifting_data_rz) + len(Above_shoulder_level_lifting_data_lz)

total_time_Above_Shoulder_level_lifting = Above_Shoulder_level_lifting / 6

Above_Shoulder_level_lifting_freq = total_time_Above_Shoulder_level_lifting / total_time * 100

print('Above_Shoulder_level_lifting_freq', round(Above_Shoulder_level_lifting_freq, 2))
# print(Above_shoulder_level_lifting_data_rz)
# print(Above_shoulder_level_lifting_data_lz)

# ####################################################

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.1]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.1]
walking_y = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) > 0.1]

# On Shoulder Carry
Shoulder_Carry_data_x = walking_x.query(
    '(`Left Shoulder Abduction/Adduction` >= 30 | `Right Shoulder Abduction/Adduction` >= 30)'
    '& (`Left Hand z` - `Left Shoulder z` >= -0.05 | `Right Hand z` - `Right Shoulder z` >= -0.05)')

Shoulder_Carry_data_y = walking_y.query(
    '(`Left Shoulder Abduction/Adduction` >= 30 | `Right Shoulder Abduction/Adduction` >= 30)'
    '& (`Left Hand z` - `Left Shoulder z` >= -0.05 | `Right Hand z` - `Right Shoulder z` >= -0.05)')

Shoulder_Carry = len(Shoulder_Carry_data_x) + len(Shoulder_Carry_data_y)

total_time_Shoulder_Carry = Shoulder_Carry / 6

Shoulder_Carry_freq = total_time_Shoulder_Carry / total_time * 100

print('Shoulder_Carry_freq', round(Shoulder_Carry_freq, 2))
# print(Shoulder_Carry_data_x)
# print(Shoulder_Carry_data_y)

# ####################################################

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.1]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.1]
walking_y = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) > 0.1]

# Side Carry Right hand
Side_Carry_Right_Hand_data_x = walking_x.query(
    '(`Right Shoulder Abduction/Adduction` >= 30) &'
    '(`Left Hand z` - `Left Shoulder z` < -0.05 & `Right Hand z` - `Right Shoulder z` < -0.05) ')

Side_Carry_Right_Hand_data_y = walking_y.query(
    '(`Right Shoulder Abduction/Adduction` >= 30) &'
    '(`Left Hand z` - `Left Shoulder z` < -0.05 & `Right Hand z` - `Right Shoulder z` < -0.05) ')

Side_Carry_Right_Hand = len(Side_Carry_Right_Hand_data_x) + len(Side_Carry_Right_Hand_data_y)

total_time_Side_Carry_Right_Hand = Side_Carry_Right_Hand / 6

Side_Carry_Right_Hand_freq = total_time_Side_Carry_Right_Hand / total_time * 100

print('Side_Carry_Right_Hand_freq', round(Side_Carry_Right_Hand_freq, 2))
# print(Side_Carry_Right_Hand_data_x)
# print(Side_Carry_Right_Hand_data_y)

# # ####################################################

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.1]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.1]
walking_y = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) > 0.1]

# Side Carry Left hand
Side_Carry_Left_Hand_data_x = walking_x.query(
    '`Left Shoulder Abduction/Adduction` >= 30 &'
    '(`Left Hand z` - `Left Shoulder z` < -0.05 & `Right Hand z` - `Right Shoulder z` < -0.05) ')

Side_Carry_Left_Hand_data_y = walking_y.query(
    '`Left Shoulder Abduction/Adduction` >= 30 &'
    '(`Left Hand z` - `Left Shoulder z` < -0.05 & `Right Hand z` - `Right Shoulder z` < -0.05) ')

Side_Carry_Left_Hand = len(Side_Carry_Left_Hand_data_x) + len(Side_Carry_Left_Hand_data_y)

total_time_Side_Carry_Left_Hand = Side_Carry_Left_Hand / 6

Side_Carry_Left_Hand_freq = total_time_Side_Carry_Left_Hand / total_time * 100

print('Side_Carry_Left_Hand_freq', round(Side_Carry_Left_Hand_freq, 2))
# print(Side_Carry_Left_Hand_data_x)
# print(Side_Carry_Left_Hand_data_y)

# # ####################################################

# Side carry total
walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.1]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.1]
walking_y = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) > 0.1]

# Side Carry Total hand
Side_Carry_Total_Hand_data_x = walking_x.query(
    '(`Left Shoulder Abduction/Adduction` >= 30 | `Right Shoulder Abduction/Adduction` >= 30) &'
    '(`Left Hand z` - `Left Shoulder z` < -0.05 & `Right Hand z` - `Right Shoulder z` < -0.05) ')

Side_Carry_Total_Hand_data_y = walking_y.query(
    '(`Left Shoulder Abduction/Adduction` >= 30 | `Right Shoulder Abduction/Adduction` >= 30) &'
    '(`Left Hand z` - `Left Shoulder z` < -0.05 & `Right Hand z` - `Right Shoulder z` < -0.05) ')

Side_Carry_Total_Hand = len(Side_Carry_Total_Hand_data_x) + len(Side_Carry_Total_Hand_data_y)

total_time_Side_Carry_Total_Hand = Side_Carry_Total_Hand / 6

Side_Carry_Total_freq = total_time_Side_Carry_Total_Hand / total_time * 100

print("Side_Carry_Total_freq", round(Side_Carry_Total_freq, 2))
# print(Side_Carry_Total_Hand_data_x)
# print(Side_Carry_Total_Hand_data_y)

# #################################################

# Carry front

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.05]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.05]
walking_y = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) > 0.05]

Carry_front_data_x = walking_x.query(
    '(`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20 |'
    '`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.25 | abs(`Right Hand x` - `Pelvis x`) >= 0.25 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.25 | abs(`Right Hand y` - `Pelvis y`) >= 0.25) ')

Carry_front_data_y = walking_y.query(
    '(`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20 |'
    '`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.25 | abs(`Right Hand x` - `Pelvis x`) >= 0.25 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.25 | abs(`Right Hand y` - `Pelvis y`) >= 0.25) ')

Carry_front = len(Carry_front_data_x) + len(Carry_front_data_y)

total_time_Carry_front = Carry_front / 6

Carry_front_freq = total_time_Carry_front / total_time * 100

print('Carry_front_freq', round(Carry_front_freq, 2))
# print(Carry_front_data_x.to_string())
# print(Carry_front_data_y.to_string())

# # ####################################################

# Pushing

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.05]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.05]
not_walking = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) <= 0.05]

Pushing_data = not_walking.query(
    '`Vertical_Pelvis Flexion/Extension` >= 5 & '
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '((`Right Hand z` - `Right Lower Leg z`) > 0 & (`Left Hand z` - `Left Lower Leg z`) > 0) &'
    '(abs(`Pelvis x` - `Head x`) >= 0.15 | abs(`Pelvis y` - `Head y`) >= 0.15) &'
    '(`Right Hip Flexion/Extension` >= 5 | `Left Hip Flexion/Extension` >= 5) &'
    '(`Right Lower Leg z` > 0.2 & `Left Lower Leg z` > 0.2) &'
    '(`Right Wrist Flexion/Extension` <= -35 & `Left Wrist Flexion/Extension` <= -35)')

Pushing = len(Pushing_data)

total_time_Pushing = Pushing / 6

Pushing_freq = total_time_Pushing / total_time * 100

print('Pushing_freq', round(Pushing_freq, 2))
# print(Pushing_data)

# ####################################################

# Pulling

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.05]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.05]
not_walking = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) <= 0.05]

Pulling_data = not_walking.query(
    '(`Vertical_Pelvis Flexion/Extension` <= -10 | `Vertical_Pelvis Flexion/Extension` >= 10) &'
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(abs(`Pelvis z` - `Right Lower Leg z`) > 0.15 & abs(`Pelvis z` - `Left Lower Leg z`) > 0.15) &'
    '(`Right Hip Flexion/Extension` >= 5 | `Left Hip Flexion/Extension` >= 5) &'
    '(`Right Wrist Flexion/Extension` >= -30 & `Left Wrist Flexion/Extension` >= -30)')

Pulling = len(Pulling_data)

total_time_Pulling = Pulling / 6

Pulling_freq = total_time_Pulling / total_time * 100

print('Pulling_freq', round(Pulling_freq, 2))
# print(Pulling_data.to_string())

# ####################################################

# Crouching
Crouching_data = data.query(
    '(abs(`Pelvis z` - `Right Lower Leg z`) <= 0.15 & abs(`Pelvis z` - `Left Lower Leg z`) <= 0.15) &'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20) &'
    '(`Vertical_Pelvis Flexion/Extension` >= 20) &'
    '((`Right Hand z` - `Right Lower Leg z`) <= 0 | (`Left Hand z` - `Left Lower Leg z`) <= 0) &'
    '(`Right Ankle Dorsiflexion/Plantarflexion` >= 10 | `Left Ankle Dorsiflexion/Plantarflexion` >= 10)')

Crouching = len(Crouching_data)

total_time_Crouching = Crouching / 6

Crouching_freq = total_time_Crouching / total_time * 100

print('Crouching_freq', round(Crouching_freq, 2))
# print(Crouching_data)

# ####################################################

# Sitting
Sitting_data = data.query(
    '(abs(`Pelvis z` - `Right Lower Leg z`) <= 0.15 & abs(`Pelvis z` - `Left Lower Leg z`) <= 0.15) &'
    '(`Right Ankle Dorsiflexion/Plantarflexion` < 15 & `Left Ankle Dorsiflexion/Plantarflexion` < 15)')

Sitting = len(Sitting_data)

total_time_Sitting = Sitting / 6

Sitting_freq = total_time_Sitting / total_time * 100

print('Sitting_freq', round(Sitting_freq, 2))
# print(Sitting_data)

# ####################################################

# Walking

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.10]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.10]
walking_y = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) > 0.10]

walking = len(walking_x) + len(walking_y)

total_time_walking = walking / 6
walking_freq = total_time_walking / total_time * 100

print('walking_freq', round(walking_freq, 2))
# print(walking_x)
# print(walking_y)

# ####################################################

# Climbing Stairs

right_walking_x = data[abs(data['Right Toe x'] - data['Right Toe x'].shift()).round(2) > 0.01]
right_not_walking_x = data[abs(data['Right Toe x'] - data['Right Toe x'].shift()).round(2) <= 0.01]
right_walking_y = right_not_walking_x[
    abs(right_not_walking_x['Right Toe y'] - right_not_walking_x['Right Toe y'].shift()).round(2) > 0.01]

left_walking_x = data[abs(data['Left Toe x'] - data['Left Toe x'].shift()).round(2) > 0.01]
left_not_walking_x = data[abs(data['Left Toe x'] - data['Left Toe x'].shift()).round(2) <= 0.01]
left_walking_y = left_not_walking_x[
    abs(left_not_walking_x['Left Toe y'] - left_not_walking_x['Left Toe y'].shift()).round(2) > 0.01]

Climbing_Stairs_data_x_right = right_walking_x[
    (abs(right_walking_x['Right Toe z'] - right_walking_x['Right Toe z'].shift()) >= 0.1)]

Climbing_Stairs_data_y_right = right_walking_y[
    (abs(right_walking_y['Right Toe z'] - right_walking_y['Right Toe z'].shift()) >= 0.1)]

Climbing_Stairs_data_x_left = left_walking_x[
    (abs(left_walking_x['Left Toe z'] - left_walking_x['Left Toe z'].shift()) >= 0.1)]

Climbing_Stairs_data_y_left = left_walking_y[
    (abs(left_walking_y['Left Toe z'] - left_walking_y['Left Toe z'].shift()) >= 0.1)]

Climbing_Stairs_data_1_right = Climbing_Stairs_data_x_right.query(
    '(20 >`Right Shoulder Flexion/Extension` >= 10 | 20 >`Left Shoulder Flexion/Extension` >= 10) |'
    '(`Right Shoulder Flexion/Extension` <= -1 | `Left Shoulder Flexion/Extension` <= -1)')

Climbing_Stairs_data_1_left = Climbing_Stairs_data_x_left.query(
    '(20 >`Right Shoulder Flexion/Extension` >= 10 | 20 >`Left Shoulder Flexion/Extension` >= 10) |'
    '(`Right Shoulder Flexion/Extension` <= -1 | `Left Shoulder Flexion/Extension` <= -1)')

Climbing_Stairs_data_2_right = Climbing_Stairs_data_y_right.query(
    '(20 >`Right Shoulder Flexion/Extension` >= 10 | 20 >`Left Shoulder Flexion/Extension` >= 10) |'
    '(`Right Shoulder Flexion/Extension` <= -1 | `Left Shoulder Flexion/Extension` <= -1)')

Climbing_Stairs_data_2_left = Climbing_Stairs_data_y_left.query(
    '(20 >`Right Shoulder Flexion/Extension` >= 10 | 20 >`Left Shoulder Flexion/Extension` >= 10) |'
    '(`Right Shoulder Flexion/Extension` <= -1 | `Left Shoulder Flexion/Extension` <= -1)')

Climbing_Stairs = len(Climbing_Stairs_data_1_left) + len(Climbing_Stairs_data_1_right) + len(
    Climbing_Stairs_data_2_left) + len(Climbing_Stairs_data_2_right)

total_time_Climbing_Stairs = Climbing_Stairs / 6

Climbing_Stairs_freq = total_time_Climbing_Stairs / total_time * 100

print('Climbing_Stairs_freq', round(Climbing_Stairs_freq, 2))

# ####################################################

# Climbing Ladder

Climbing_Ladder_data = data[(abs(data['Right Toe z'] - data['Right Toe z'].shift()) >= 0.1) |
                            (abs(data['Left Toe z'] - data['Left Toe z'].shift()) >= 0.1)]
Climbing_Ladder_data_1 = Climbing_Ladder_data.query(
    '(`Right Shoulder Flexion/Extension` >= 20 & `Left Shoulder Flexion/Extension` >= 20)')

Climbing_Ladder = len(Climbing_Ladder_data_1)

total_time_Climbing_Ladder = Climbing_Ladder / 6

Climbing_Ladder_freq = total_time_Climbing_Ladder / total_time * 100

print('Climbing_Ladder_freq', round(Climbing_Ladder_freq, 2))
# print(Climbing_Ladder_data_1)

# ####################################################

# Climbing Stools
Climbing_Stools_data = data[(abs(data['Right Toe z'] - data['Right Toe z'].shift()) >= 0.1) |
                            (abs(data['Left Toe z'] - data['Left Toe z'].shift()) >= 0.1)]

Climbing_Stools = len(Climbing_Stools_data)

total_time_Climbing_Stools = Climbing_Stools / 6

Climbing_Stools_freq = total_time_Climbing_Stools / total_time * 100

print('Climbing_Stools_freq', round(Climbing_Stools_freq, 2))
# print(Climbing_Stools_data.to_string())

# ####################################################

# Kneeling
Kneeling_data = data.query(
    '`Right Lower Leg z` <= 0.15| `Left Lower Leg z` <= 0.15')

Kneeling = len(Kneeling_data)

total_time_Kneeling = Kneeling / 6

Kneeling_freq = total_time_Kneeling / total_time * 100

print('Kneeling_freq', round(Kneeling_freq, 2))
# print(Kneeling_data)

# ####################################################

# Trunk Forward Bending
trunk_forward_bending_data = data[data['Vertical_Pelvis Flexion/Extension'] >= 20]
trunk_forward_bending = len(trunk_forward_bending_data)

total_time_trunk_forward_bending = trunk_forward_bending / 6

trunk_forward_bending_freq = total_time_trunk_forward_bending / total_time * 100

print('trunk_forward_bending_frequency', round(trunk_forward_bending_freq, 2))
# print(trunk_forward_bending_data)

# Trunk Backward Bending
trunk_backward_bending_data = data[data['Vertical_Pelvis Flexion/Extension'] <= -20]
trunk_backward_bending = len(trunk_backward_bending_data)

total_time_trunk_backward_bending = trunk_backward_bending / 6

trunk_backward_bending_freq = total_time_trunk_backward_bending / total_time * 100

print('trunk_backward_bending_frequency', round(trunk_backward_bending_freq, 2))
# print(trunk_backward_bending_data)

# Trunk Rotation
trunk_rotation_data = data.query('`L5S1 Axial Bending` >= 10 | `L5S1 Axial Bending` <= -10')

trunk_rotation = len(trunk_rotation_data)

total_time_trunk_rotation = trunk_rotation / 6

trunk_rotation_freq = total_time_trunk_rotation / total_time * 100

print('trunk_rotation_frequency', round(trunk_rotation_freq, 2))
# print(trunk_rotation_data)

# ####################################################

# Below Shoulder Level Reaching
Below_shoulder_level_Reaching_data = data.query(
    '((`Vertical_Pelvis Flexion/Extension` < 45) &'
    '(`Right Hand z` < `Right Shoulder z`- @B/2 |'
    ' `Left Hand z` < `Left Shoulder z`- @B/2) &'
    '(45 > `Right Shoulder Flexion/Extension` >= 20 | 45 > `Left Shoulder Flexion/Extension` >= 20) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) )|'
    '((`Vertical_Pelvis Flexion/Extension` >= 45) &'
    '(abs(`Left Hand x` - `Left Shoulder x`) > 0.1 | abs(`Right Hand x` - `Right Shoulder x`) > 0.1 |'
    'abs(`Left Hand y` - `Left Shoulder y`) > 0.1 | abs(`Right Hand y` - `Right Shoulder y`) > 0.1) &'
    '(45 > `Right Shoulder Flexion/Extension` >= 20 | 45 > `Left Shoulder Flexion/Extension` >= 20) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) )')

Below_shoulder_level_Reaching = len(Below_shoulder_level_Reaching_data)

total_time_Below_shoulder_level_Reaching = Below_shoulder_level_Reaching / 6

Below_shoulder_level_Reaching_freq = total_time_Below_shoulder_level_Reaching / total_time * 100

print('Below_shoulder_level_Reaching_freq', round(Below_shoulder_level_Reaching_freq, 2))
# print(Below_shoulder_level_Reaching_data)

# ####################################################

# Shoulder Level Reaching
shoulder_level_Reaching_data = data.query(
    '((`Vertical_Pelvis Flexion/Extension` < 45) &'
    '(`Right Shoulder z`- @B/2 <= `Right Hand z` < `Right Shoulder z` + 0.1 |'
    ' `Left Shoulder z`- @B/2 <= `Left Hand z` < `Left Shoulder z` + 0.1) &'
    '(`Right Shoulder Flexion/Extension` >= 45 | `Left Shoulder Flexion/Extension` >= 45) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30))|'
    '((`Vertical_Pelvis Flexion/Extension` >= 45) &'
    '(abs(`Left Hand x` - `Left Shoulder x`) <= 0.1 | abs(`Right Hand x` - `Right Shoulder x`) <= 0.1 |'
    'abs(`Left Hand y` - `Left Shoulder y`) <= 0.1 | abs(`Right Hand y` - `Right Shoulder y`) <= 0.1) &'
    '(`Right Shoulder Flexion/Extension` >= 45 | `Left Shoulder Flexion/Extension` >= 45) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30))')

Shoulder_level_Reaching = len(shoulder_level_Reaching_data)

total_time_Shoulder_level_Reaching = Shoulder_level_Reaching / 6

Shoulder_level_Reaching_freq = total_time_Shoulder_level_Reaching / total_time * 100

print('Shoulder_level_Reaching_freq', round(Shoulder_level_Reaching_freq, 2))

# #################################################

# Above Shoulder Level Reaching
Above_shoulder_level_Reaching_data = data.query(
    '((`Vertical_Pelvis Flexion/Extension` < 45) &'
    '(`Right Hand z`>=`Right Shoulder z` + 0.1 | `Left Hand z`>= `Left Shoulder z` + 0.1) &' 
    '(`Right Shoulder Flexion/Extension` >= 45 | `Left Shoulder Flexion/Extension` >= 45) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30))|'
    '((`Vertical_Pelvis Flexion/Extension` >= 45) &'
    '(abs(`Left Hand x` - `Left Shoulder x`) > 0.1 | abs(`Right Hand x` - `Right Shoulder x`) > 0.1 |'
    'abs(`Left Hand y` - `Left Shoulder y`) > 0.1 | abs(`Right Hand y` - `Right Shoulder y`) > 0.1) &'
    '(`Right Shoulder Flexion/Extension` >= 45 | `Left Shoulder Flexion/Extension` >= 45) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30))')

Above_shoulder_level_Reaching = len(Above_shoulder_level_Reaching_data)

total_time_Above_shoulder_level_Reaching = Above_shoulder_level_Reaching / 6

Above_shoulder_level_Reaching_freq = total_time_Above_shoulder_level_Reaching / total_time * 100

print('Above_shoulder_level_Reaching_freq', round(Above_shoulder_level_Reaching_freq, 2))
# print(Above_shoulder_level_Reaching_data)

# ####################################################

# Sideway Shoulder Reaching

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.01]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.01]
not_walking = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) <= 0.01]

Sideway_Shoulder_Reaching_data = not_walking.query(
    '(`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) &'
    '((`Right Shoulder z` - 0.3) <= `Right Hand z` | (`Left Shoulder z` - 0.3) <= `Left Hand z`) &'
    '(`Right Shoulder Abduction/Adduction` >= 30 | `Left Shoulder Abduction/Adduction` >= 30) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) &'
    '(`T8_Head Axial Bending` >= 10 | `T8_Head Axial Bending` <= -10) ')

Sideway_Shoulder_Reaching = len(Sideway_Shoulder_Reaching_data)

total_time_Sideway_Shoulder_Reaching = Sideway_Shoulder_Reaching / 6

Sideway_Shoulder_Reaching_freq = total_time_Sideway_Shoulder_Reaching / total_time * 100

print('Sideway_Shoulder_Reaching_freq', round(Sideway_Shoulder_Reaching_freq, 2))
# print(Sideway_Shoulder_Reaching_data)

# ################################################### #

# Behind Shoulder Reaching
Behind_Shoulder_Reaching_data = data.query(
    '(`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) &'
    '(`Right Shoulder Flexion/Extension` <= -5 | `Left Shoulder Flexion/Extension` <= -5) &'
    '(abs(`Left Hand x` - `Pelvis x`) >= 0.30 | abs(`Right Hand x` - `Pelvis x`) >= 0.30 |'
    'abs(`Left Hand y` - `Pelvis y`) >= 0.30 | abs(`Right Hand y` - `Pelvis y`) >= 0.30) &'
    '(`T8_Head Axial Bending` >= 10 | `T8_Head Axial Bending` <= -10) &'
    '(`L5S1 Axial Bending` >= 5 | `L5S1 Axial Bending` <= -5)')

Behind_Shoulder_Reaching = len(Behind_Shoulder_Reaching_data)

total_time_Behind_Shoulder_Reaching = Behind_Shoulder_Reaching / 6

Behind_Shoulder_Reaching_freq = total_time_Behind_Shoulder_Reaching / total_time * 100

print('Behind_Shoulder_Reaching_freq', round(Behind_Shoulder_Reaching_freq, 2))
# print(Behind_Shoulder_Reaching_data)

# ####################################################

# Neck Forward Bending
neck_forward_bending_data = data[data['T8_Head Flexion/Extension'] >= 20]

neck_forward_bending = len(neck_forward_bending_data)

total_time_neck_forward_bending = neck_forward_bending / 6

neck_forward_bending_freq = total_time_neck_forward_bending / total_time * 100

print('neck_forward_bending_freq', round(neck_forward_bending_freq, 2))
# print(neck_forward_bending_data)

# Neck Backward Bending
neck_backward_bending_data = data[data['T8_Head Flexion/Extension'] <= -20]

neck_backward_bending = len(neck_backward_bending_data)

total_time_neck_backward_bending = neck_backward_bending / 6

neck_backward_bending_freq = total_time_neck_backward_bending / total_time * 100

print('neck_backward_bending_freq', round(neck_backward_bending_freq, 2))
# print(neck_backward_bending_data)

# Neck Twist Tilt
neck_twist_tilt_data = data.query(
    '(`T8_Head Lateral Bending` >= 20 | `T8_Head Lateral Bending` <= -20) | '
    '(`T8_Head Axial Bending` >= 20 | `T8_Head Axial Bending` <= -20)')

neck_twist_tilt = len(neck_twist_tilt_data)

total_time_neck_twist_tilt = neck_twist_tilt / 6

neck_twist_tilt_freq = total_time_neck_twist_tilt / total_time * 100

print('neck_twist_tilt_freq', round(neck_twist_tilt_freq, 2))
# print(neck_twist_tilt_freq)

# ##################################################

# Elbow Flexion Extension
Elbow_Flexion_Extension_data = data.query(
    '(`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20)')

Elbow_Flexion_Extension = len(Elbow_Flexion_Extension_data)

total_time_Elbow_Flexion_Extension = Elbow_Flexion_Extension / 6

Elbow_Flexion_Extension_freq = total_time_Elbow_Flexion_Extension / total_time * 100

print('Elbow_Flexion_Extension_freq', round(Elbow_Flexion_Extension_freq, 2))
# print(Elbow_Flexion_Extension_data)

# ####################################################

# Wrist Flexion Extension
Wrist_Flexion_Extension_data = data.query(
    '(`Right Wrist Flexion/Extension` >= 15 | `Left Wrist Flexion/Extension` >= 15) |'
    '(`Right Wrist Flexion/Extension` <= -15 | `Left Wrist Flexion/Extension` <= -15)')

Wrist_Flexion_Extension = len(Wrist_Flexion_Extension_data)

total_time_Wrist_Flexion_Extension = Wrist_Flexion_Extension / 6

Wrist_Flexion_Extension_freq = total_time_Wrist_Flexion_Extension / total_time * 100

print('Wrist_Flexion_Extension_freq', round(Wrist_Flexion_Extension_freq, 2))
# print(Wrist_Flexion_Extension_data)

# ####################################################

# Wrist Bending
Wrist_Bending_data = data.query(
    '(`Right Wrist Ulnar Deviation/Radial Deviation` >= 10 | `Left Wrist Ulnar Deviation/Radial Deviation` >= 10) |'
    '(`Right Wrist Ulnar Deviation/Radial Deviation` <= -15 | `Left Wrist Ulnar Deviation/Radial Deviation` <= -15)')

Wrist_Bending = len(Wrist_Bending_data)

total_time_Wrist_Bending = Wrist_Bending / 6

Wrist_Bending_freq = total_time_Wrist_Bending / total_time * 100

print('Wrist_Bending_freq', round(Wrist_Bending_freq, 2))
# print(Wrist_Bending_data)

# ####################################################

# Wrist Rotate
Wrist_Rotate_data = data.query(
    '(`Right Wrist Pronation/Supination` >= 15 | `Left Wrist Pronation/Supination` >= 15) | '
    '(`Right Wrist Pronation/Supination` <= -15 | `Left Wrist Pronation/Supination` <= -15)')

Wrist_Rotate = len(Wrist_Rotate_data)
total_time_Wrist_Rotate = Wrist_Rotate / 6

Wrist_Rotate_freq = total_time_Wrist_Rotate / total_time * 100

print('Wrist_Rotate_freq', round(Wrist_Rotate_freq, 2))
# print(Wrist_Rotate_data)

# ####################################################

# Ankle Flexion Extension
Ankle_Flexion_Extension_data = data.query(
    '(`Right Ankle Dorsiflexion/Plantarflexion` >= 15 | `Left Ankle Dorsiflexion/Plantarflexion` >= 15) |'
    '(`Right Ankle Dorsiflexion/Plantarflexion` <= -15 | `Left Ankle Dorsiflexion/Plantarflexion` <= -15)')

Ankle_Flexion_Extension = len(Ankle_Flexion_Extension_data)

total_time_Ankle_Flexion_Extension = Ankle_Flexion_Extension / 6

Ankle_Flexion_Extension_freq = total_time_Ankle_Flexion_Extension / total_time * 100

print('Ankle_Flexion_Extension_freq', round(Ankle_Flexion_Extension_freq, 2))
# print(Ankle_Flexion_Extension_data)

# ####################################################

# Ankle Rotate
Ankle_Rotate_data = data.query(
    '(`Right Ankle Internal/External Rotation` >= 15 | `Left Ankle Internal/External Rotation` >= 15) | '
    '(`Right Ankle Internal/External Rotation` <= -15 | `Left Ankle Internal/External Rotation` <= -15)')

Ankle_Rotate = len(Ankle_Rotate_data)

total_time_Ankle_Rotate = Ankle_Rotate / 6

Ankle_Rotate_freq = total_time_Ankle_Rotate / total_time * 100

print('Ankle_Rotate_freq', round(Ankle_Rotate_freq, 2))
# print(Ankle_Rotate_data)

# ####################################################

# Dynamic Pushing

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.05]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.05]
walking_y = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) > 0.05]

Dynamic_Pushing_data_x = walking_x.query(
    '`Vertical_Pelvis Flexion/Extension` >= 10 & '
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(abs(`Pelvis x` - `Head x`) >= 0.15 | abs(`Pelvis y` - `Head y`) >= 0.15) &'
    '(`Right Hip Flexion/Extension` >= 5 & `Left Hip Flexion/Extension` >= 5) &'
    '(`Right Wrist Flexion/Extension` <= -35 & `Left Wrist Flexion/Extension` <= -35)')

Dynamic_Pushing_data_y = walking_y.query(
    '`Vertical_Pelvis Flexion/Extension` >= 10 & '
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(abs(`Pelvis x` - `Head x`) >= 0.15 | abs(`Pelvis y` - `Head y`) >= 0.15) &'
    '(`Right Hip Flexion/Extension` >= 5 & `Left Hip Flexion/Extension` >= 5) &'
    '(`Right Wrist Flexion/Extension` <= -35 & `Left Wrist Flexion/Extension` <= -35)')

Dynamic_Pushing = len(Dynamic_Pushing_data_x) + len(Dynamic_Pushing_data_y)

total_time_Dynamic_Pushing = Dynamic_Pushing / 6

Dynamic_Pushing_freq = total_time_Dynamic_Pushing / total_time * 100

print('Dynamic_Pushing_freq', round(Dynamic_Pushing_freq, 2))
# print(Dynamic_Pushing_data_x)
# print(Dynamic_Pushing_data_y)

# ####################################################

# Dynamic Pulling

walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) > 0.1]
not_walking_x = data[abs(data['Pelvis x'] - data['Pelvis x'].shift()).round(2) <= 0.1]
walking_y = not_walking_x[abs(not_walking_x['Pelvis y'] - not_walking_x['Pelvis y'].shift()).round(2) > 0.1]

Dynamic_Pulling_data_x = walking_x.query(
    '(`Vertical_Pelvis Flexion/Extension` <= -5 | `Vertical_Pelvis Flexion/Extension` >= 10) &'
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(`Right Hip Flexion/Extension` >= 5 & `Left Hip Flexion/Extension` >= 5) &'
    '(`Right Wrist Flexion/Extension` >= -30 & `Left Wrist Flexion/Extension` >= -30)')

Dynamic_Pulling_data_y = walking_y.query(
    '(`Vertical_Pelvis Flexion/Extension` <= -5 | `Vertical_Pelvis Flexion/Extension` >= 10) &'
    '((`Right Elbow Flexion/Extension` >= 20 | `Left Elbow Flexion/Extension` >= 20) |'
    '(`Right Shoulder Flexion/Extension` >= 20 | `Left Shoulder Flexion/Extension` >= 20)) &'
    '(`Right Hip Flexion/Extension` >= 5 & `Left Hip Flexion/Extension` >= 5) &'
    '(`Right Wrist Flexion/Extension` >= -30 & `Left Wrist Flexion/Extension` >= -30)')

Dynamic_Pulling = len(Dynamic_Pulling_data_x) + len(Dynamic_Pulling_data_y)

total_time_Dynamic_Pulling = Dynamic_Pulling / 6

Dynamic_Pulling_freq = total_time_Dynamic_Pulling / total_time * 100

print('Dynamic_Pulling_freq', round(Dynamic_Pulling_freq, 2))
# print(Dynamic_Pulling_data_x)
# print(Dynamic_Pulling_data_y)

# ####################################################

# Crawling
Crawling_data = data.query(
    '((abs(`Head z` - `Right Foot z`) <= 0.1 | abs(`Head z` - `Left Foot z`) <= 0.1) &'
    '(abs(`Pelvis z` - `Right Foot z`) <= 0.1 | abs(`Pelvis z` - `Left Foot z`) <= 0.1)) |'
    '(abs(`Head z` - `Pelvis z`) <= 0.1 &'
    'abs(`Right Hand z` - `Right Foot z`) <= 0.1 & abs(`Left Hand z` - `Left Foot z`) <= 0.1 &'
    'abs(`Right Lower Leg z` - `Right Foot z`) <= 0.1 & abs(`Left Lower Leg z` - `Left Lower Leg z`) <= 0.1)')

Crawling = len(Crawling_data)

total_time_Crawling = Crawling / 6

Crawling_freq = total_time_Crawling / total_time * 100

print('Crawling_freq', round(Crawling_freq, 2))
# print(Crawling_data)

# ####################################################

# Knee Flexion Extension
Knee_Flexion_Extension_data = data.query(
    '(`Right Knee Flexion/Extension` >= 70 | `Left Knee Flexion/Extension` >= 70)')

Knee_Flexion_Extension = len(Knee_Flexion_Extension_data) - len(Sitting_data)

total_time_Knee_Flexion_Extension = Knee_Flexion_Extension / 6

Knee_Flexion_Extension_freq = total_time_Knee_Flexion_Extension / total_time * 100

print('Knee_Flexion_Extension_freq', round(Knee_Flexion_Extension_freq, 2))
# print(Knee_Flexion_Extension_data)

# ####################################################

# Squatting
squatting_data = data.query(
    '(abs(`Pelvis z` - `Right Lower Leg z`) <= 0.15 & abs(`Pelvis z` - `Left Lower Leg z`) <= 0.15) &'
    '(`Vertical_Pelvis Flexion/Extension` >= 10) &'
    '(`Right Ankle Dorsiflexion/Plantarflexion` < 15 | `Left Ankle Dorsiflexion/Plantarflexion` < 15) &'
    '(`Right Hand z` > `Right Lower Leg z` | `Left Hand z` > `Left Lower Leg z`)')

squatting = len(Knee_Flexion_Extension_data) - len(Sitting_data)

total_time_squatting = squatting / 6

squatting_freq = total_time_squatting / total_time * 100

print('squatting_freq', round(Knee_Flexion_Extension_freq, 2))
# print(squatting_data)


# ####################################################

# Standing

walking_x_list = list(walking_x.Frame)
walking_y_list = list(walking_y.Frame)
climbing_stairs_list_1 = list(Climbing_Stairs_data_1_right.Frame)
climbing_stairs_list_2 = list(Climbing_Stairs_data_2_right.Frame)
climbing_stairs_list_3 = list(Climbing_Stairs_data_1_left.Frame)
climbing_stairs_list_4 = list(Climbing_Stairs_data_2_left.Frame)
climbing_ladder_list = list(Climbing_Ladder_data_1.Frame)
climbing_stools_list = list(Climbing_Stools_data.Frame)
sitting_list = list(Sitting_data.Frame)
crouching_list = list(Crouching_data.Frame)
kneeling_list = list(Kneeling_data.Frame)
low_level_lifting_list_1 = list(low_level_lifting_data_rz.Frame)
low_level_lifting_list_2 = list(low_level_lifting_data_lz.Frame)
Knee_level_lifting_list_1 = list(Knee_level_lifting_data_rz.Frame)
Knee_level_lifting_list_2 = list(Knee_level_lifting_data_lz.Frame)
Pulling_list_1 = list(Pulling_data.Frame)
Pushing_list_1 = list(Pushing_data.Frame)
Dynamic_Pushing_list_1 = list(Dynamic_Pushing_data_x.Frame)
Dynamic_Pulling_list_1 = list(Dynamic_Pushing_data_x.Frame)
Dynamic_Pushing_list_2 = list(Dynamic_Pushing_data_y.Frame)
Dynamic_Pulling_list_2 = list(Dynamic_Pushing_data_y.Frame)
Crawling_list = list(Crawling_data.Frame)
squatting_list = list(Knee_Flexion_Extension_data.Frame)
trunk_forward_bending_list = list(trunk_forward_bending_data.Frame)
trunk_backward_bending_list = list(trunk_backward_bending_data.Frame)

standing = len(data) - len(set(walking_x_list + walking_y_list + climbing_stairs_list_1 + climbing_stairs_list_2 +
                               climbing_stairs_list_3 + climbing_stairs_list_4 + climbing_ladder_list + climbing_stools_list +
                               sitting_list + crouching_list + kneeling_list + low_level_lifting_list_1 + low_level_lifting_list_2 +
                               Knee_level_lifting_list_1 + Knee_level_lifting_list_2 + Pulling_list_1 + Pushing_list_1 +
                               Dynamic_Pushing_list_1 + Dynamic_Pulling_list_1 + Dynamic_Pushing_list_2 + Dynamic_Pulling_list_2 +
                               Crawling_list + squatting_list + trunk_forward_bending_list + trunk_backward_bending_list))


total_time_standing = standing / 6

standing_freq = total_time_standing / total_time * 100

print('standing_freq', round(standing_freq, 2))

# ####################################################

# Write to excel:

wb = load_workbook("result.xlsx")
sheet = wb.worksheets[0]


def write_to_excel(row, percent):
    sheet = wb.worksheets[0]
    if 0 <= percent and percent < 2:
        sheet.cell(row=row, column=3).value = percent.round(2)
    elif 2 <= percent and percent <= 5:
        sheet.cell(row=row, column=4).value = percent.round(2)
    elif 5 < percent and percent <= 33:
        sheet.cell(row=row, column=5).value = percent.round(2)
    elif 33 < percent and percent <= 66:
        sheet.cell(row=row, column=6).value = percent.round(2)
    elif 66 < percent and percent <= 100:
        sheet.cell(row=row, column=7).value = percent.round(2)


write_to_excel(4, low_level_lifting_freq)
write_to_excel(5, Knee_level_lifting_freq)
write_to_excel(6, waist_level_lifting_freq)
write_to_excel(7, Shoulder_level_lifting_freq)
write_to_excel(8, Above_Shoulder_level_lifting_freq)
write_to_excel(10, Carry_front_freq)
write_to_excel(11, Side_Carry_Total_freq)
write_to_excel(12, Side_Carry_Right_Hand_freq)
write_to_excel(13, Side_Carry_Left_Hand_freq)
write_to_excel(14, Shoulder_Carry_freq)
write_to_excel(16, Pushing_freq)
write_to_excel(17, Pulling_freq)
write_to_excel(18, Dynamic_Pushing_freq)
write_to_excel(19, Dynamic_Pulling_freq)
write_to_excel(23, Sitting_freq)
write_to_excel(24, standing_freq)
write_to_excel(25, walking_freq)
write_to_excel(26, Climbing_Stairs_freq)
write_to_excel(27, Climbing_Ladder_freq)
write_to_excel(28, Climbing_Stools_freq)
write_to_excel(29, Crouching_freq)
write_to_excel(30, squatting_freq)
write_to_excel(31, Kneeling_freq)
write_to_excel(32, Crawling_freq)
write_to_excel(34, trunk_forward_bending_freq)
write_to_excel(35, trunk_rotation_freq)
write_to_excel(36, trunk_backward_bending_freq)
write_to_excel(38, Below_shoulder_level_Reaching_freq)
write_to_excel(39, Shoulder_level_Reaching_freq)
write_to_excel(40, Above_shoulder_level_Reaching_freq)
write_to_excel(41, Sideway_Shoulder_Reaching_freq)
write_to_excel(42, Behind_Shoulder_Reaching_freq)
write_to_excel(44, neck_forward_bending_freq)
write_to_excel(45, neck_backward_bending_freq)
write_to_excel(46, neck_twist_tilt_freq)
write_to_excel(48, Elbow_Flexion_Extension_freq)
write_to_excel(50, Wrist_Flexion_Extension_freq)
write_to_excel(51, Wrist_Bending_freq)
write_to_excel(52, Wrist_Rotate_freq)
write_to_excel(54, Ankle_Flexion_Extension_freq)
write_to_excel(55, Ankle_Rotate_freq)

wb.save("result.xlsx")
